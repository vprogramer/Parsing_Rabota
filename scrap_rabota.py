import re
import requests
from bs4 import BeautifulSoup
import openpyxl
import os.path


def write_in_excel(func):
    def wrapper(*args):
        file = "rabota.xlsx"
        # file = "work.xlsx"
        if os.path.exists(file):
            wb = openpyxl.load_workbook(file)
        else:
            wb = openpyxl.Workbook()
        # sheet_catalog = wb.sheetnames
        new_sheet = None

        while True:
            answer = input("Do you want to create new sheet for new data? Type y or n \n")
            if answer == "y":
                new_sheet = create_new_sheet(wb)
                break
            elif answer == "n":
                break
            else:
                continue

        if bool(new_sheet) is True:
            sheet = wb[new_sheet]
        else:
            sheet = wb.active

        info_rabota = func(*args)
        for info in info_rabota:
            sheet.append(info)
            wb.save(file)
        sheet.append((" ", " ", " "))
        wb.save(file)
    return wrapper


def create_new_sheet(wb):
    sheetname = input("Type sheet name: \n")
    sheet = wb.create_sheet(sheetname)
    sheet.append(('Технология', 'Вакансии', 'Резюме', 'чел/место', 'Перспективность'))
    return sheetname


def get_number_of_vacancies(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                                 'Chrome/93.0.4577.82 Safari/537.36',}
        r = requests.get(url, headers=headers)
        soup = BeautifulSoup(r.text, "lxml")

        num_vac = soup.find("h1", class_="bloko-header-section-3")
        # print(num_vac.text)
        return int(re.search(r"\d+", num_vac.text).group())
    except AttributeError:
        return 0


def get_number_of_resume(url):
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                                 'Chrome/93.0.4577.82 Safari/537.36',}
        r = requests.get(url, headers=headers)
        soup = BeautifulSoup(r.text, "lxml")

        num_resume = soup.find("h1", {"class": "bloko-header-1", "data-qa": "bloko-header-1"}).find_next()
        # print(num_resume.text)
        return int(re.search(r"\d+", num_resume.text).group())
    except AttributeError:
        return 0


@write_in_excel
def get_info_about_vacancies_and_resume(language, technologies, *experience):
    sheet = list()

    for tech in technologies.split(", "):
        vacancy = 0
        resume = 0
        for exp in experience:
            # Перебрать в цікле навыкі і страніцы і опыт работы
            url_vacancy = "https://rabota.by/search/vacancy?clusters=true&area=1002&ored_clusters=true{0}" \
                          "&enable_snippets=true&salary=&text={1}+{2}&" \
                          "from=suggest_post".format(exp, language, tech)

            url_resume = "https://rabota.by/search/resume?clusters=True&area=1002&currency_code=BYR&ored_clusters=True" \
                         "&order_by=relevance&logic=normal&pos=full_text&exp_period=all_time{0}" \
                         "&text={1}+{2}".format(exp, language, tech)
            vacancy += get_number_of_vacancies(url_vacancy)
            resume += get_number_of_resume(url_resume)

        sheet.append((tech, vacancy, resume))
    return sheet


def main():
    technologies = "Git, Django, DRF, MySQL, PostgreSQL, redis, asyncio, Flask, aiohttp, FastApi, Starlette, Sanic, " \
                   "unit+тесты, MongoDB, Celery, Telegram, bs4, beautifulsoup, selenium, pydantic, asyncio, " \
                   " Falcon, threading, multiprocessing, tornado"
    technologies_ml = "OpenCV, ML, Machine+Learning, Computer+Vision, sklearn, matplotlib, Keras, TensorFlow, " \
                      "PyTorch, numpy, Databricks, Spark, Kafka, Airflow, Apache AirFlow, SciPy, scikit-image, " \
                      "torchvision, Caffe"
    technologies_devops = "Linux, CI/CD, NGINX, Docker, k8s, Nomad, OpenShift, Bash, Ansible, Jenkins, AWS, Git, " \
                          "Azure, Docker+Compose, Grafana, Prometheus, Clickhouse, Kubernetes, Terraform, " \
                          "PowerShell, OSI, TCP/IP, Zabbix, Kafka, Spark, Redis, Cassandra, ClickHouse, " \
                          "ElasticSearch, Kibana, Fluentd, Zipkin, php-fpm, Memcached, RabbitMQ, ELK, Nexus, " \
                          "Sonarqube, Hashicorp+Nomad, Buffers, protobuf, Apache, gitlab, haproxy, grpc, python"

    work_types = "developer, web, ml, data+science, maсhine+learning, big+data, Analyst, Data+Engineer, ETL, BI, " \
                 "backend, back-end, Automation+QA, devops, security, secops"

    language = "Python"

    experiences = ("&experience=noExperience", "&experience=between1And3", "&experience=between3And6",
                  "&experience=moreThan6", "&experience=")

    # lang = ["Python", "C++", "Java", "C#", "Kotlin", "Go", "Dart", "Ruby", "JavaScript", "PHP", "Swift", "Scala"]


    get_info_about_vacancies_and_resume(language, technologies, experiences[0], experiences[1], experiences[2], experiences[3])


if __name__ == "__main__":
    main()
